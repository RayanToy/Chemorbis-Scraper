"""Excel output formatting module."""

import logging

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)


def format_output_excel(file_path: str, sheet_name: str = "Sheet1") -> None:
    """Apply Excel table formatting and number formats to output file.

    Creates a styled Excel table with appropriate number formats:
    - Columns A-D, G: Text format
    - Column E: Integer number format
    - Columns F, K: Date format (DD.MM.YYYY)
    - Columns H-J: Integer number format

    Args:
        file_path: Path to the Excel file to format.
        sheet_name: Name of the worksheet to format.
    """
    logger.info(f"Formatting Excel file: {file_path}")

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    max_col = ws.max_column
    max_row = ws.max_row

    # Apply cell formats
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        # Columns A-D: Text
        for cell in row[0:4]:
            cell.number_format = "@"

        # Column E (Price): Integer
        row[4].number_format = "0"

        # Column F (Date): Short date
        row[5].number_format = "DD.MM.YYYY"

        # Column G (Agency): Text
        row[6].number_format = "@"

        # Columns H-J (Month, Year, Week): Integer
        for cell in row[7:10]:
            cell.number_format = "0"

        # Column K (Week Start Date): Short date
        if len(row) > 10:
            row[10].number_format = "DD.MM.YYYY"

    # Create Excel table
    table = Table(
        displayName="ChemOrbisData",
        ref=f"A1:K{max_row}",
    )

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(file_path)
    logger.info(f"Excel formatting applied: {file_path}")